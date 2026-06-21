# -*- coding: utf-8 -*-
"""Genera casos_de_prueba.xlsx — plan de testing del TP5 Playa (Grupo 13)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Casos de Prueba"

ACC = "264653"; HDR = "2A9D8F"; ZEBRA = "EAF3F2"
CAT = "DCE0E8"
thin = Side(style="thin", color="BBBBBB")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

# Título e instrucciones
ws.merge_cells("A1:H1")
ws["A1"] = "TP5 — Playa de Estacionamiento (Grupo 13) — Plan de Pruebas"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=ACC)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:H2")
ws["A2"] = ("Completá 'Resultado obtenido', '¿OK?' y 'Observaciones'. Para cada caso, configurá los "
            "parámetros indicados, tocá ▶ Simular y compará con el resultado esperado. Después mandame el archivo.")
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

# Encabezados (fila 4)
headers = ["ID", "Categoría", "Caso de prueba", "Cómo configurarlo / parámetros",
           "Resultado esperado", "Resultado obtenido", "¿OK?", "Observaciones"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(4, c, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HDR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borde
ws.row_dimensions[4].height = 24

casos = [
    # (categoría, caso, params, esperado)
    ("Parámetros", "Valores por defecto", "N=8, media=13, tmáx=24, i vacío",
     "Simula cientos de filas, la playa se llena (LL), recaudación y % util > 0, sin errores."),
    ("Parámetros", "N = 1 lugar", "N=1, tmáx=24",
     "Un solo sector: mucha cola/abandono; % utilización alto."),
    ("Parámetros", "N = 10 (pregunta b)", "N=10, tmáx=24",
     "Más recaudación y menos abandonos que con 8; % util más bajo."),
    ("Parámetros", "N = 20", "N=20, tmáx=24",
     "Casi sin abandonos; % utilización bajo."),
    ("Parámetros", "Media de llegada baja", "media=5, N=8",
     "Llegan más autos → más cola y más abandonos."),
    ("Parámetros", "Media de llegada alta", "media=30, N=8",
     "Llegan pocos autos → casi sin cola; % util bajo."),
    ("Parámetros", "Tiempo máx corto", "tmáx=2, i vacío",
     "Pocas filas; el reloj final no pasa de 2 h."),
    ("Parámetros", "Tiempo máx largo", "tmáx=72",
     "Muchas filas; corta por tiempo o por 100.000 iteraciones."),

    ("Euler", "Subir T", "T=10 (mirar tab Auditoría Euler)",
     "El término 0.2·T crece; el cobro llega antes al umbral (tiempo menor)."),
    ("Euler", "T = 0", "T=0",
     "dD/dt = C + t²; cobros un poco más largos."),
    ("Euler", "h = 0.5", "h=0.5",
     "Tabla de Euler con menos filas; tiempo de cobro algo menos preciso."),
    ("Euler", "h = 0.1 (default)", "h=0.1",
     "Tabla detallada; con C=0: chico/util ≈ 7.4 min, grande ≈ 8.1 min."),
    ("Euler", "Umbral por tipo", "ver resumen Euler",
     "D=180 para Grande; D=130 para Pequeño y Utilitario."),
    ("Euler", "C afecta el cobro", "buscar un cobro con C>0",
     "A mayor C, el cobro llega antes al umbral (menor tiempo)."),

    ("Precios", "Cambiar precios", "Pequeño=1000, Grande=2000, Utilitario=4000",
     "El importe de cada auto y la recaudación cambian proporcionalmente."),
    ("Precios", "Verificar importe", "mirar columna Importe",
     "Importe = horas × precio del tipo (ej. Grande 3h a 1500 = 4500)."),

    ("Prob. tipo", "Todos Pequeño", "%Peq=100, %Gra=0, %Uti=0",
     "Todos los autos son Pequeño; recaudación más baja."),
    ("Prob. tipo", "Todos Utilitario", "%Peq=0, %Gra=0, %Uti=100",
     "Todos Utilitario; recaudación más alta."),
    ("Prob. tipo", "Otra proporción", "%Peq=20, %Gra=20, %Uti=60",
     "Predominan utilitarios; verificar proporción aproximada en la tabla."),
    ("Prob. horas", "Todos 4 horas", "%1h=0, %2h=0, %3h=0, %4h=100",
     "Todos estacionan 4 h; la playa se llena más y el % util sube."),
    ("Prob. horas", "Todos 1 hora", "%1h=100, resto 0",
     "Todos 1 h; rotan rápido, menos cola."),
    ("Prob. horas", "Uniforme", "%1h=25, %2h=25, %3h=25, %4h=25",
     "Distribución pareja de horas."),

    ("Visualización", "i vacío muestra todo", "i vacío, j=0",
     "El vector muestra TODAS las filas simuladas (no queda vacío)."),
    ("Visualización", "Limitar a 20 filas", "i=20, j=0",
     "Muestra 20 filas; la simulación se detiene en ~20 (no simula de más)."),
    ("Visualización", "Mostrar desde hora j", "j=5, i vacío",
     "Solo se muestran filas con reloj ≥ 5."),
    ("Visualización", "Re-aplicar i/j", "cambiar i a 10 y tocar 'Re-aplicar i/j'",
     "Re-filtra la vista sin volver a simular."),
    ("Visualización", "Última fila", "cualquier corrida",
     "Abajo se ve la fila del estado final (instante X)."),

    ("Lógica", "Abandono por playa llena", "media=4, N=8",
     "Con ocupados=8, el auto que llega dice '(abandona)' en rojo y no entra."),
    ("Lógica", "Liberar lugar al terminar", "ver un fin_estacionamiento",
     "Al fin_estacionamiento el lugar pasa a Libre al instante (ocupados baja)."),
    ("Lógica", "Cola de cobro y C", "buscar fin_cobro con cola ≥ 2",
     "Al entrar uno a cobrar, C = los que quedan en la cola (cola − 1)."),
    ("Lógica", "Zona de cobro capacidad 1", "mirar Servidor Cobro",
     "Solo 1 auto en cobro a la vez; los demás esperan en la cola."),
    ("Lógica", "Auto en cola NO ocupa lugar", "fila con cola ≥ 1",
     "Los autos EsperandoCobro no se cuentan en 'ocupados'."),

    ("Estadísticas", "Recaudación (pregunta a)", "default",
     "Recaudación = Σ (horas × precio) de los autos que pagaron. Validar a mano una muestra."),
    ("Estadísticas", "% utilización (pregunta c)", "default",
     "% util = área de ocupación / (N × tiempo simulado) × 100."),
    ("Estadísticas", "What-if 10 vs 8 (pregunta b)", "correr N=8 y N=10",
     "Comparar recaudación: con 10 lugares ≥ con 8."),

    ("Interfaz", "Columnas congeladas", "hacer scroll horizontal",
     "#, Reloj y Evento quedan fijas a la izquierda."),
    ("Interfaz", "Scroll vertical sincronizado", "hacer scroll vertical",
     "Las 3 columnas fijas scrollean junto con el resto."),
    ("Interfaz", "Copiar a Excel", "botón 'Copiar a Excel' y pegar en Excel",
     "Pega encabezados + filas visibles, alineado en columnas."),
    ("Interfaz", "Salto a Euler", "click en una celda 'T. Cobro (h)'",
     "Cambia al tab Auditoría Euler en la iteración de ese auto."),
    ("Interfaz", "Resumen Euler clickable", "click en una fila del resumen",
     "Salta al detalle de esa integración."),
    ("Interfaz", "Tablas Euler ajustables", "arrastrar el divisor del medio",
     "Se redimensionan las dos tablas; el divisor se pinta azul al pasar el mouse."),
    ("Interfaz", "Tema claro sin grilla", "vista general",
     "Fondo claro (sin blanco), tablas sin líneas internas, solo el borde."),
    ("Interfaz", "Colores de estado", "mirar columnas de estado y tipo",
     "Libre verde, Ocupado naranja, tipos coloreados, '(abandona)' en rojo."),
]

dv = DataValidation(type="list", formula1='"OK,Falla,Parcial"', allow_blank=True)
ws.add_data_validation(dv)

fila = 5
cat_actual = None
for i, (cat, caso, params, esp) in enumerate(casos, 1):
    ws.cell(fila, 1, i)
    ws.cell(fila, 2, cat)
    ws.cell(fila, 3, caso)
    ws.cell(fila, 4, params)
    ws.cell(fila, 5, esp)
    # 6,7,8 vacías para completar
    zebra = ZEBRA if i % 2 == 0 else "FFFFFF"
    for c in range(1, 9):
        cell = ws.cell(fila, c)
        cell.border = borde
        cell.alignment = Alignment(vertical="top", wrap_text=True,
                                   horizontal="center" if c in (1, 7) else "left")
        cell.fill = PatternFill("solid", fgColor=CAT if c == 2 else zebra)
        if c == 2:
            cell.font = Font(bold=True, size=9)
    dv.add(ws.cell(fila, 7))
    ws.row_dimensions[fila].height = 30
    fila += 1

# Anchos
anchos = {"A": 5, "B": 13, "C": 26, "D": 30, "E": 46, "F": 28, "G": 9, "H": 24}
for col, w in anchos.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A5"
wb.save("casos_de_prueba.xlsx")
print("Generado: casos_de_prueba.xlsx con", len(casos), "casos")
